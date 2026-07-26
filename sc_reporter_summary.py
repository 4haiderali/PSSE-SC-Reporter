# ============================================================
# PSS/E Short Circuit Reporter (Summary Only)
# Batch IEC 60909 fault analysis to a formatted Excel summary
#
# Version : 1.0.3
# Author  : Haider Ali (github.com/4haiderali)
# License : MIT
#
# Usage:
#   1. Place this script in a folder containing (or above) your
#      PSS/E .sav case file(s). It will recursively search for
#      .sav files in the current folder and subfolders.
#   2. Edit the USER SETTINGS section below - at minimum SC_BUSES
#      and PSSE_PATH.
#   3. Run from the PSS/E Python 2.7 environment:
#        exec(open("sc_reporter_summary.py").read())
#      or launch via the PSS/E script runner.
#   4. Results are written per-case to the output folder defined
#      by OUTPUT_FOLDER_NAME, as one Excel file per .sav case.
#
# Requirements:
#   PSS/E v34, Python 2.7, openpyxl
#   Install openpyxl in the PSS/E Python 2.7 environment:
#     <PSSE_PATH>\python.exe -m pip install openpyxl
#
# Changelog:
#   1.0.0  Initial release. Recursive .sav discovery, IEC 60909 short
#          circuit analysis on user-defined buses, formatted Excel summary
#          output (bus, name, kV, 3-phase/1-phase kA), a metadata footer,
#          and safe-save logic for a locked output file.
#   1.0.1  All IEC 60909 fault settings (psspy.iecs_4 STATUS/VALUES) are
#          now named, individually documented USER SETTINGS instead of a
#          single hardcoded array. USER SETTINGS moved to the top of the
#          file, ahead of PSS/E path setup.
#   1.0.2  Fixed a silent-empty-output bug: the parser and Excel columns
#          were hardcoded for 3-phase + LG faults (5 numbers per report
#          line). Since 1.0.1 made fault types configurable, any other
#          combination (e.g. 3-phase only) produced a report with a
#          different number of columns that silently matched nothing,
#          writing an empty workbook with no error. The parser and output
#          columns are now built dynamically from the enabled fault types.
#          Also added a guard: only REPORT_OPTION = 0 (summary table) is
#          supported by this parser.
#   1.0.3  Parsing and output columns are now derived from the report's own
#          header (detect_report_fault_types), not from the requested
#          settings. PSS/E can silently compute fewer fault types than
#          requested when a case is missing negative/zero-sequence data
#          (same reason LG/LLG/LL gray out in the IEC dialog) - previously
#          this produced another silent empty output. A mismatch between
#          requested and actual fault types is now flagged both in the
#          console and as a highlighted warning row in the Excel footer.
# ============================================================

from __future__ import print_function, division  # MUST be the first import
import sys
import os
import re
import time
from datetime import datetime

# --- OpenPyXL import (compatible across versions) ---
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        from openpyxl.cell import get_column_letter
except ImportError:
    print("ERROR: openpyxl module not found. Install it with:")
    print("  <PSSE_PATH>\\python.exe -m pip install openpyxl")
    sys.exit()

__version__ = "1.0.3"
TOOL_NAME = "PSS/E Short Circuit Reporter (Summary Only)"
TOOL_AUTHOR = "Haider Ali"
TOOL_GITHUB = "github.com/4haiderali"

# ============================================================
# USER SETTINGS
# ============================================================

# REQUIRED: set the bus number(s) to fault before running.
# The script will stop with a clear error if this is left empty.
# Example: SC_BUSES = [101, 205, 30012]
SC_BUSES = []

# PSS/E 34 Python installation. Adjust to match your install.
# Common alternatives:
#   r"C:\Program Files\PTI\PSSE34\PSSPY27"       (64-bit installer)
#   r"C:\Program Files (x86)\PTI\PSSE34\PSSPY27" (32-bit / default)
PSSE_PATH = r"C:\Program Files (x86)\PTI\PSSE34\PSSPY27"

# Folder search depth (0 = current folder, 1 = one level down, etc.)
SEARCH_LEVELS = [0, 1, 2, 3]

# Output folder name (created alongside this script).
OUTPUT_FOLDER_NAME = "Short Circuit Reports"

# Fixed-width bus name field in the PSS/E SC report (v34 caps names at 12).
BUS_NAME_FIELD_WIDTH = 12

# Sanity ceiling for parsed kV; raise if your model uses higher voltages.
MAX_PLAUSIBLE_KV = 800.0

# --- IEC 60909 fault settings (psspy.iecs_4 STATUS/VALUES arrays) ---
# See PSS/E API docs, IECS_4, for the full meaning of each option.

# Fault types to include
INCLUDE_3PHASE_FAULT = True     # STATUS(1)
INCLUDE_LG_FAULT = True         # STATUS(2) line-to-ground
INCLUDE_LLG_FAULT = False       # STATUS(3) line-line-to-ground
INCLUDE_LL_FAULT = False        # STATUS(4) line-to-line

# Report format
REPORT_OPTION = 0               # STATUS(5): 0=summary table, 1=total currents,
                                 #   2=contributions to N levels, 3=both
CONTRIBUTION_LEVELS = 0         # STATUS(6): used only if REPORT_OPTION is 2 or 3

# Fault location
FAULT_LOCATION = 0              # STATUS(7): 0=network bus, 1=PSU LV bus,
                                 #   2=auxiliary transformer LV bus

# Additional fault scenarios
INCLUDE_LINE_OUT_FAULTS = False # STATUS(8)
INCLUDE_LINE_END_FAULTS = False # STATUS(9)

# Network modeling options
TAP_OPTION = 0                  # STATUS(10): 0=unchanged, 1=taps 1.0pu & shift 0deg,
                                 #   2=taps 1.0pu only, 3=shift 0deg only
LINE_CHARGING_OPTION = 1        # STATUS(11): 0=unchanged, 1=zero in pos/neg seq,
                                 #   2=zero in all sequences
SHUNT_OPTION = 1                # STATUS(12): same scale as LINE_CHARGING_OPTION
DC_LINE_OPTION = 0              # STATUS(13): 0=blocked, 1=represent as load
ZERO_SEQ_CORRECTION = 0         # STATUS(14): 0=ignore, 1=apply xfmr zero-seq correction
VOLTAGE_FACTOR_OPTION = 0       # STATUS(15): 0=max fault C, 1=min fault C,
                                 #   2=user C for max, 3=user C for min
LOAD_OPTION = 1                 # STATUS(16): 0=unchanged, 1=zero in pos/neg seq,
                                 #   2=zero in all sequences
GENERATOR_REACTANCE_OPTION = 0  # STATUS(17): 0=subtransient, 1=transient, 2=synchronous

# Fault calculation values
BREAKER_CONTACT_TIME = 0.1      # VALUES(1), seconds
USER_VOLTAGE_FACTOR_C = 1.0     # VALUES(2), used only if VOLTAGE_FACTOR_OPTION is 2 or 3

# --- Excel styling ---
BORDER_COLOR = "2F75B5"               # Strong Blue
HEADER_INSIDE_BORDER_COLOR = "EDF5FC" # Very Light Blue
HEADER_FILL_COLOR = "4A85B8"          # Standard Blue
LIGHT_FILL_COLOR = "DDEBF7"           # Light Blue
TEXT_COLOR = "FFFFFF"                 # White

if not SC_BUSES:
    raise RuntimeError(
        "SC_BUSES is empty. Please define SC_BUSES = [bus numbers to fault] "
        "in the USER SETTINGS section before running."
    )

FAULT_STATUS = [
    1 if INCLUDE_3PHASE_FAULT else 0,
    1 if INCLUDE_LG_FAULT else 0,
    1 if INCLUDE_LLG_FAULT else 0,
    1 if INCLUDE_LL_FAULT else 0,
    REPORT_OPTION,
    CONTRIBUTION_LEVELS,
    FAULT_LOCATION,
    1 if INCLUDE_LINE_OUT_FAULTS else 0,
    1 if INCLUDE_LINE_END_FAULTS else 0,
    TAP_OPTION,
    LINE_CHARGING_OPTION,
    SHUNT_OPTION,
    DC_LINE_OPTION,
    ZERO_SEQ_CORRECTION,
    VOLTAGE_FACTOR_OPTION,
    LOAD_OPTION,
    GENERATOR_REACTANCE_OPTION,
]
FAULT_VALUES = [BREAKER_CONTACT_TIME, USER_VOLTAGE_FACTOR_C]

# Columns in the SC report follow this fixed order (matches STATUS(1)-(4)):
# 3-phase, then LG, then LLG, then LL - only enabled types get a column pair.
# This drives what's requested from PSS/E. It is NOT necessarily what ends up
# in the report: PSS/E can silently compute fewer fault types than requested
# if a case is missing negative/zero-sequence data (the same reason the IEC
# 60909 dialog grays out LG/LLG/LL faults). See detect_report_fault_types().
_ALL_FAULT_TYPES = [
    ("3PH", "3-Phase", INCLUDE_3PHASE_FAULT),
    ("LG", "1-Phase (LG)", INCLUDE_LG_FAULT),
    ("LLG", "LLG", INCLUDE_LLG_FAULT),
    ("LL", "LL", INCLUDE_LL_FAULT),
]
REQUESTED_FAULT_TYPES = [(k, lbl) for k, lbl, on in _ALL_FAULT_TYPES if on]

if not REQUESTED_FAULT_TYPES:
    raise RuntimeError(
        "No fault type is enabled. Set at least one of INCLUDE_3PHASE_FAULT, "
        "INCLUDE_LG_FAULT, INCLUDE_LLG_FAULT, INCLUDE_LL_FAULT to True."
    )

if REPORT_OPTION != 0:
    raise RuntimeError(
        "REPORT_OPTION = {} is not supported. This script parses the fault "
        "current summary table only (REPORT_OPTION = 0). Contribution and "
        "total-fault-current report layouts use a different structure.".format(REPORT_OPTION)
    )

# ============================================================
# PSS/E v34 / Python 2.7 setup
# ============================================================

# Try common PSS/E 34 locations if the configured path is not found.
if not os.path.isdir(PSSE_PATH):
    _common_psse_paths = [
        r"C:\Program Files (x86)\PTI\PSSE34\PSSPY27",
        r"C:\Program Files\PTI\PSSE34\PSSPY27",
    ]
    _found = None
    for _p in _common_psse_paths:
        if os.path.isdir(_p):
            _found = _p
            break
    if _found is not None:
        PSSE_PATH = _found
    else:
        raise RuntimeError(
            "PSS/E Python path not found: {}\n"
            "Edit PSSE_PATH in the USER SETTINGS section to match "
            "your PSS/E 34 installation.".format(PSSE_PATH)
        )

if PSSE_PATH not in sys.path:
    sys.path.append(PSSE_PATH)
os.environ["PATH"] = PSSE_PATH + ";" + os.environ.get("PATH", "")

try:
    import psse34
except ImportError:
    pass

import psspy
import redirect

redirect.psse2py()
psspy.psseinit(50000)

# ============================================================
# System initialization
# ============================================================

mydir = os.getcwd()
os.chdir(mydir)
print("Working directory: {}".format(mydir))

output_folder = os.path.join(mydir, OUTPUT_FOLDER_NAME)
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# ============================================================
# Helper functions
# ============================================================

def find_sav_files(root_dir, allowed_levels):
    savs = []
    for root, dirs, files in os.walk(root_dir):
        rel_path = os.path.relpath(root, root_dir)
        depth = 0 if rel_path == "." else rel_path.count(os.sep) + 1
        if depth in allowed_levels:
            for f in files:
                if f.lower().endswith(".sav"):
                    savs.append(os.path.join(root, f))
    return savs


def detect_report_fault_types(lines):
    """
    Determine which fault-type columns are actually present in this report,
    by reading its header - independent of what was requested. PSS/E can
    compute fewer fault types than requested if a case is missing negative
    or zero-sequence data (the same reason LG/LLG/LL are grayed out in the
    IEC 60909 Fault Calculation dialog when sequence data isn't available).
    """
    header_text = "\n".join(lines[:40])
    detected = []
    if "THREE PHASE FAULT" in header_text:
        detected.append(("3PH", "3-Phase"))
    if re.search(r'(?<!L)LG FAULT', header_text):
        detected.append(("LG", "1-Phase (LG)"))
    if "LLG FAULT" in header_text:
        detected.append(("LLG", "LLG"))
    if "LL FAULT" in header_text:
        detected.append(("LL", "LL"))
    return detected


def filter_existing_buses(bus_list):
    """Return only buses that exist in the currently loaded case."""
    existing = []
    for b in bus_list:
        try:
            ierr, _ = psspy.busdat(int(b), 'BASE')
            if ierr == 0:
                existing.append(int(b))
        except Exception:
            pass
    return existing


def run_iec_sc_analysis(sav_path, report_path, log_path):
    """Run IEC 60909 short circuit analysis on SC_BUSES for one case."""
    ierr = psspy.case(sav_path)
    if ierr != 0:
        print("Error loading {}".format(sav_path))
        return False

    alert_path = os.path.splitext(report_path)[0] + "_Alert.txt"
    prompt_path = os.path.splitext(report_path)[0] + "_Prompt.txt"

    psspy.report_output(2, report_path, [0, 0])
    psspy.progress_output(2, log_path, [0, 0])
    psspy.alert_output(2, alert_path, [0, 0])
    psspy.prompt_output(2, prompt_path, [0, 0])

    try:
        psspy.short_circuit_units(1)
        psspy.short_circuit_z_units(1)
        psspy.short_circuit_coordinates(1)
        psspy.short_circuit_z_coordinates(1)

        valid_buses = filter_existing_buses(SC_BUSES)
        if not valid_buses:
            print("  - No valid SC buses found in this case. Skipping.")
            return False

        psspy.bsys(0, 0, [0.0, 0.0], 0, [], len(valid_buses), valid_buses, 0, [], 0, [])

        psspy.iecs_4(0, 0, FAULT_STATUS, FAULT_VALUES, "", "", "")
    except Exception as e:
        print("Error during simulation: {}".format(str(e)))
        return False
    finally:
        psspy.report_output(1, "", [0, 0])
        psspy.progress_output(1, "", [0, 0])
        psspy.alert_output(1, "", [0, 0])
        psspy.prompt_output(1, "", [0, 0])

    return True


def format_excel_sheet(ws, case_title, active_types):
    """Apply header/title styling, borders, widths, and row heights.
    active_types is the list of fault types actually detected in this
    report (see detect_report_fault_types), not the requested settings."""
    title_font = Font(bold=True, color=TEXT_COLOR, size=16)
    header_font = Font(bold=True, color=TEXT_COLOR, size=14)
    data_font = Font(bold=False, color="000000", size=12)

    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_FILL_COLOR, end_color=LIGHT_FILL_COLOR, fill_type="solid")

    center_style = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin_border = Border(left=Side(style='thin', color=BORDER_COLOR),
                          right=Side(style='thin', color=BORDER_COLOR),
                          top=Side(style='thin', color=BORDER_COLOR),
                          bottom=Side(style='thin', color=BORDER_COLOR))

    light_inside = Side(style="thin", color=HEADER_INSIDE_BORDER_COLOR)

    max_col = 4 + len(active_types)
    last_col_letter = get_column_letter(max_col)

    line1, line2 = psspy.titldt()
    ws['A1'] = "Short Circuit Fault Currents - " + line1.title()
    ws.merge_cells('A1:{}1'.format(last_col_letter))

    ws['A2'] = "S. No."
    ws['B2'] = "Bus Number"
    ws['C2'] = "Bus Name"
    ws['D2'] = "Base kV"

    for i, (_key, label) in enumerate(active_types):
        col_letter = get_column_letter(5 + i)
        ws['{}2'.format(col_letter)] = label
        ws['{}3'.format(col_letter)] = "(kA)"

    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:B3')
    ws.merge_cells('C2:C3')
    ws.merge_cells('D2:D3')

    identifiers = [""] + [chr(ord('A') + i) for i in range(max_col - 1)]
    identifiers = ["" if s == "" else "({})".format(s) for s in identifiers]
    for i, char in enumerate(identifiers, 1):
        ws.cell(row=4, column=i).value = char

    max_row = ws.max_row

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = title_font
                cell.alignment = center_style
            elif cell.row in (2, 3):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_style
            elif cell.row == 4:
                cell.fill = light_fill
                cell.font = data_font
                cell.alignment = center_style
            else:
                cell.font = data_font
                if cell.col_idx == 1:
                    cell.fill = light_fill
                    cell.alignment = center_style
                elif cell.col_idx == 3:
                    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                else:
                    cell.alignment = center_style

    header_max_row = 3
    for r in ws.iter_rows(min_row=1, max_row=header_max_row, max_col=max_col):
        for cell in r:
            b = cell.border
            r_idx, c_idx = cell.row, cell.column
            left = light_inside if c_idx > 1 else b.left
            right = light_inside if c_idx < max_col else b.right
            top = light_inside if r_idx > 1 else b.top
            bottom = light_inside if r_idx < header_max_row else b.bottom
            if r_idx == 4:
                bottom = Side(style='thin', color=BORDER_COLOR)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    widths = {'A': 10, 'B': 17, 'C': 25, 'D': 15}
    for i in range(len(active_types)):
        widths[get_column_letter(5 + i)] = 17
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 25

    ws.freeze_panes = "A5"


def add_metadata_footer(ws, case_title, data_end_row, max_col, warning=None):
    """Write tool name, version, author, GitHub, case, and timestamp below
    the table. If warning is set (e.g. a requested-vs-actual fault-type
    mismatch), it's added as a highlighted row so it's visible even if the
    console output wasn't watched during a batch run."""
    label_font = Font(bold=True, color="808080", size=9, italic=True)
    value_font = Font(bold=False, color="808080", size=9, italic=True)
    warning_font = Font(bold=True, color="C00000", size=10, italic=True)

    start_row = data_end_row + 2  # one blank row gap after the table

    footer_rows = [
        ("Generated by", "{} v{}".format(TOOL_NAME, __version__)),
        ("Author", TOOL_AUTHOR),
        ("GitHub", TOOL_GITHUB),
        ("Case", case_title),
        ("Generated on", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for i, (label, value) in enumerate(footer_rows):
        r = start_row + i
        c_label = ws.cell(row=r, column=1, value=label)
        c_label.font = label_font
        c_value = ws.cell(row=r, column=2, value=value)
        c_value.font = value_font
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    if warning:
        r = start_row + len(footer_rows) + 1  # one blank row gap
        c = ws.cell(row=r, column=1, value="WARNING: " + warning)
        c.font = warning_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)


def parse_sc_report_to_excel(report_path, excel_path, case_title):
    print("Formatting Excel: {}".format(excel_path))

    if not os.path.exists(report_path):
        return

    with open(report_path, 'r') as f:
        lines = f.readlines()

    active_types = detect_report_fault_types(lines)
    if not active_types:
        print("  - WARNING: could not detect any fault-type columns in the "
              "report header for {}. Skipping.".format(case_title))
        return

    requested_keys = [k for k, _ in REQUESTED_FAULT_TYPES]
    actual_keys = [k for k, _ in active_types]
    missing_keys = [k for k in requested_keys if k not in actual_keys]

    seq_warning = None
    if missing_keys:
        seq_warning = (
            "Requested {} but report only contains {} - likely missing "
            "negative/zero-sequence data for this case.".format(
                ", ".join(requested_keys), ", ".join(actual_keys))
        )
        print("  - WARNING: {}".format(seq_warning))

    # Prefix: bus number, [name+kV], unit. Then one (magnitude, angle) pair
    # per fault type actually present in this report, then voltage factor.
    pattern_parts = [
        r"^\s*(\d+)\s+",      # Bus Number
        r"\[(.*?)\]\s+",      # Name and kV
        r"(\w+)\s+",          # Unit
    ]
    pattern_parts += [r"([-\d\.]+)\s+([-\d\.]+)\s+"] * len(active_types)
    pattern_parts.append(r"([-\d\.]+)")  # Voltage factor
    pattern = re.compile("".join(pattern_parts))

    data_rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        match = pattern.search(line)
        if not match:
            continue

        bus_num = int(match.group(1))
        name_kv_raw = match.group(2)  # unstripped - slice depends on position

        name_part = name_kv_raw[:BUS_NAME_FIELD_WIDTH]
        kv_part = name_kv_raw[BUS_NAME_FIELD_WIDTH:]

        bus_name = name_part.strip()
        kv_str = kv_part.strip()

        try:
            base_kv_val = float(kv_str)
            if base_kv_val <= 0 or base_kv_val > MAX_PLAUSIBLE_KV:
                raise ValueError("implausible kV: {}".format(base_kv_val))
        except ValueError:
            base_kv_val = ""
            bus_name = name_kv_raw.strip()
            print("  - WARNING: could not parse kV for bus {} from '{}'; "
                  "check BUS_NAME_FIELD_WIDTH.".format(bus_num, name_kv_raw))

        row_data = {'num': bus_num, 'name': bus_name, 'kv': base_kv_val}
        for i, (key, _label) in enumerate(active_types):
            mag_group = 4 + i * 2  # magnitude group index for this fault type
            row_data[key] = float(match.group(mag_group)) / 1000.0

        data_rows.append(row_data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Short Circuit Summary"

    current_row = 5
    for idx, row_data in enumerate(data_rows, 1):
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=row_data['num'])
        ws.cell(row=current_row, column=3, value=row_data['name'])
        ws.cell(row=current_row, column=4, value=row_data['kv'])

        for i, (key, _label) in enumerate(active_types):
            col = 5 + i
            val = row_data[key]
            c = ws.cell(row=current_row, column=col, value=val)
            if val > 0 and round(val, 2) != 0:
                c.number_format = '0.00'

        current_row += 1

    last_data_row = current_row - 1
    format_excel_sheet(ws, case_title, active_types)
    add_metadata_footer(ws, case_title, last_data_row, 4 + len(active_types), seq_warning)

    # --- Safe save: auto-close the target file if it's open in Excel ---
    try:
        import win32com.client
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            xl = None

        abs_target = os.path.abspath(excel_path).lower()

        if xl is not None:
            try:
                open_books = [wb_open for wb_open in xl.Workbooks]
            except Exception:
                open_books = []
            for wb_open in open_books:
                try:
                    if wb_open.FullName and wb_open.FullName.lower() == abs_target:
                        wb_open.Close(SaveChanges=False)
                        break
                except Exception:
                    pass

        for _ in range(10):
            if not os.path.exists(excel_path):
                break
            try:
                os.remove(excel_path)
                break
            except Exception:
                time.sleep(0.2)
    except Exception:
        pass

    wb.active = wb.index(ws)
    wb.save(excel_path)
    print("Saved Excel: {}".format(excel_path))


# ============================================================
# Main execution
# ============================================================

def main():
    print("{} v{} - {}".format(TOOL_NAME, __version__, TOOL_GITHUB))
    print("Starting Short Circuit Batch Processing (Python 2.7)...")

    sav_files = find_sav_files(mydir, SEARCH_LEVELS)
    if not sav_files:
        print("No .sav files found.")
        return

    print("Found {} .sav file(s).".format(len(sav_files)))

    text_output_folder = os.path.join(output_folder, "Text Logs")
    if not os.path.exists(text_output_folder):
        os.makedirs(text_output_folder)

    for i, sav_path in enumerate(sav_files):
        print("\nProcessing: {}".format(os.path.basename(sav_path)))

        case_name = os.path.splitext(os.path.basename(sav_path))[0]
        report_file = os.path.join(text_output_folder, "{}-{}_SC.txt".format(i, case_name))
        log_file = os.path.join(text_output_folder, "{}-{}_Log.txt".format(i, case_name))
        excel_file = os.path.join(output_folder, "{}-{}_SC.xlsx".format(i, case_name))

        for f in [report_file, log_file]:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except Exception:
                    pass

        success = run_iec_sc_analysis(sav_path, report_file, log_file)
        if success:
            print("  - Report saved.")
            parse_sc_report_to_excel(report_file, excel_file, case_name)

    print("\nProcessing Complete.")


if __name__ == "__main__":
    main()
